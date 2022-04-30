import openpyxl as xl; 
from xlsconfig import *
import os
import pyexcel as p
import pandas as pd
import numpy as np
import glob
import shutil
from datetime import date

today = date.today()
 
inputWorkSheet = inputWorkSheet
outputWorkSheet = outputWorkSheet
#global readColumn1
#global writeColumn1
def CleanUp():
    newXLS = os.listdir("./")
    for xlsxFile in newXLS:
        if xlsxFile.endswith(".xlsx") or xlsxFile.endswith(".XLSX"):
           if xlsxFile != "blogistica.xlsx":
              shutil.move(xlsxFile, "./processed/" + xlsxFile)

def WriteExcelOutput():
    
    all_data = pd.DataFrame()
    for f in glob.glob('./*.xlsx'):
       print(f)
       df = pd.read_excel(f, index = None, header = None, skiprows = 1)
       #df[12] = df[12].dt.strftime('%d/%m/%Y')
       #print(df[12])
       all_data = all_data.append(df, ignore_index=True)
    writer = pd.ExcelWriter('blogistica.xlsx', engine='xlsxwriter')
    all_data.to_excel(writer, sheet_name='Sheet1', index  = False, header = None)
    writer.save()


def ExcelActivities(inputXLS):
    #date_style = Style(number_format="M/D/YYYY")
    mydate = today.strftime("%d/%m/%Y")
    wb1 = xl.load_workbook(inputXLS) 
    ws1 = wb1.active

    wb2 = xl.load_workbook(outputFile) 
    ws2 = wb2.active 
    mr = ws1.max_row 
    mc = ws1.max_column 
    for i in range (2, mr +1): 
            
            if 'readColumn1' in globals() and 'writeColumn1' in globals():
                c1 = ws1.cell(row = i, column = readColumn1) 
                ws2.cell(row = i, column = writeColumn1).value = c1.value
                
            if 'readColumn2' in globals() and 'writeColumn2' in globals():
                c2 = ws1.cell(row = i, column = readColumn2)
                # #str(c2.value)
                # #print(type(c2.value))
                # #tf = c2.value
                # #c2_time = tf.strftime("%d/%m/%Y")
                # #print(c2.value)
                # #print(type(c2.value))
                # #c2.value.style.number_format ="M/D/YYYY"
                # #ws2.cell(row = i, column = writeColumn2).number_format="M/D/YYYY"
                # #myvalue = str(c2.value)
                # #myvalue = myvalue[0:9]
                # #print(myvalue + "wef23")
                ws2.cell(row = i, column = writeColumn2).value = c2.value
                
                #print(c2.value)
            if 'readColumn3' in globals() and 'writeColumn3' in globals():
                c3 = ws1.cell(row = i, column = readColumn3) 
                ws2.cell(row = i, column = writeColumn3).value = c3.value

            if 'readColumn4' in globals() and 'writeColumn4' in globals():    
                c4 = ws1.cell(row = i, column = readColumn4) 
                ws2.cell(row = i, column = writeColumn4).value = c4.value

            if 'readColumn5' in globals() and 'writeColumn5' in globals():
                c5 = ws1.cell(row = i, column = readColumn5) 
                ws2.cell(row = i, column = writeColumn5).value = c5.value

            if 'readColumn6' in globals() and 'writeColumn6' in globals():
                c6 = ws1.cell(row = i, column = readColumn6) 
                ws2.cell(row = i, column = writeColumn6).value = c6.value

            if 'readColumn7' in globals() and 'writeColumn7' in globals():
                c7 = ws1.cell(row = i, column = readColumn7) 
                ws2.cell(row = i, column = writeColumn7).value = c7.value

            if 'readColumn8' in globals() and 'writeColumn8' in globals():
                c8 = ws1.cell(row = i, column = readColumn8) 
                ws2.cell(row = i, column = writeColumn8).value = c8.value
         
            if 'readColumn9' in globals() and 'writeColumn9' in globals():
                c9 = ws1.cell(row = i, column = readColumn9) 
                ws2.cell(row = i, column = writeColumn9).value = c9.value
        
            if 'readColumn10' in globals() and 'writeColumn10' in globals():
                c10 = ws1.cell(row = i, column = readColumn10) 
                ws2.cell(row = i, column = writeColumn10).value = c10.value
                
            if 'readColumn11' in globals() and 'writeColumn11' in globals():
                c11 = ws1.cell(row = i, column = readColumn11) 
                ws2.cell(row = i, column = writeColumn11).value = c11.value
  
            if 'readColumn12' in globals() and 'writeColumn12' in globals():
                c12 = ws1.cell(row = i, column = readColumn12) 
                ws2.cell(row = i, column = writeColumn12).value = c12.value

            if 'readColumn13' in globals() and 'writeColumn13' in globals():
                c13 = ws1.cell(row = i, column = readColumn13) 
                ws2.cell(row = i, column = writeColumn13).value = c13.value
                
            if 'readColumn14' in globals() and 'writeColumn14' in globals():
                c14 = ws1.cell(row = i, column = readColumn14) 
                ws2.cell(row = i, column = writeColumn14).value = c14.value

            if 'readColumn15' in globals() and 'writeColumn15' in globals():
                c15 = ws1.cell(row = i, column = readColumn15) 
                ws2.cell(row = i, column = writeColumn15).value = c15.value
 
            if 'readColumn16' in globals() and 'writeColumn16' in globals():
                c16 = ws1.cell(row = i, column = readColumn16) 
                ws2.cell(row = i, column = writeColumn16).value = c16.value
            if 'readColumn17' in globals() and 'writeColumn17' in globals():
                c17 = ws1.cell(row = i, column = readColumn17) 
                ws2.cell(row = i, column = writeColumn17).value = c17.value 
            if 'readColumn18' in globals() and 'writeColumn18' in globals():
                c18 = ws1.cell(row = i, column = readColumn18) 
                ws2.cell(row = i, column = writeColumn18).value = c18.value
            if 'readColumn19' in globals() and 'writeColumn19' in globals():
                c19 = ws1.cell(row = i, column = readColumn19) 
                ws2.cell(row = i, column = writeColumn19).value = c19.value
            if 'readColumn20' in globals() and 'writeColumn20' in globals():
                c20 = ws1.cell(row = i, column = readColumn20) 
                ws2.cell(row = i, column = writeColumn20).value = c20.value
            if 'readColumn21' in globals() and 'writeColumn21' in globals():
                c21 = ws1.cell(row = i, column = readColumn21) 
                ws2.cell(row = i, column = writeColumn21).value = c21.value
            if 'readColumn22' in globals() and 'writeColumn22' in globals():
                c22 = ws1.cell(row = i, column = readColumn22) 
                ws2.cell(row = i, column = writeColumn22).value = c22.value      
            if 'readColumn23' in globals() and 'writeColumn23' in globals():
                c23 = ws1.cell(row = i, column = readColumn23) 
                ws2.cell(row = i, column = writeColumn23).value = c23.value
            if 'readColumn24' in globals() and 'writeColumn24' in globals():
                c24 = ws1.cell(row = i, column = readColumn24) 
                ws2.cell(row = i, column = writeColumn24).value = c24.value      
            if 'readColumn25' in globals() and 'writeColumn25' in globals():
                c25 = ws1.cell(row = i, column = readColumn25) 
                ws2.cell(row = i, column = writeColumn25).value = c25.value      
            if 'readColumn26' in globals() and 'writeColumn26' in globals():
                c26 = ws1.cell(row = i, column = readColumn26) 
                ws2.cell(row = i, column = writeColumn26).value = c26.value
            
            if 'readColumn27' in globals() and 'writeColumn27' in globals():
                c27 = ws1.cell(row = i, column = readColumn27) 
                ws2.cell(row = i, column = writeColumn27).value = c27.value
            
            if 'readColumn28' in globals() and 'writeColumn28' in globals():
                c28 = ws1.cell(row = i, column = readColumn28) 
                ws2.cell(row = i, column = writeColumn28).value = c28.value
            
            if 'readColumn29' in globals() and 'writeColumn29' in globals():
                c29 = ws1.cell(row = i, column = readColumn29) 
                ws2.cell(row = i, column = writeColumn29).value = c29.value      
            
            if 'readColumn30' in globals() and 'writeColumn30' in globals():
                c30 = ws1.cell(row = i, column = readColumn30) 
                ws2.cell(row = i, column = writeColumn30).value = c30.value
            
            if 'writeColumn31' in globals():
                c31 = ws1.cell(row = i, column = readColumn31) 
                ws2.cell(row = i, column = writeColumn31).value = c31.value      
            
            if 'writeColumn32' in globals():
                c32 = ws1.cell(row = i, column = readColumn32) 
                ws2.cell(row = i, column = writeColumn32).value = c32.value
            
            if 'readColumn33' in globals() and 'writeColumn33' in globals():
                c33 = ws1.cell(row = i, column = readColumn32) 
                ws2.cell(row = i, column = writeColumn33).value = c33.value
            
            if 'readColumn34' in globals() and 'writeColumn34' in globals():
                c34 = ws1.cell(row = i, column = readColumn34) 
                ws2.cell(row = i, column = writeColumn34).value = c34.value
            
            if 'readColumn35' in globals() and 'writeColumn35' in globals():
                c35 = ws1.cell(row = i, column = readColumn35) 
                ws2.cell(row = i, column = writeColumn32).value = c35.value
            
            if 'readColumn36' in globals() and 'writeColumn36' in globals():
                c36 = ws1.cell(row = i, column = readColumn36) 
                ws2.cell(row = i, column = writeColumn32).value = c36.value
            
            if 'readColumn37' in globals() and 'writeColumn37' in globals():
                c37 = ws1.cell(row = i, column = readColumn37) 
                ws2.cell(row = i, column = writeColumn32).value = c37.value
            
            if 'readColumn38' in globals() and 'writeColumn38' in globals():
                c38 = ws1.cell(row = i, column = readColumn38) 
                ws2.cell(row = i, column = writeColumn38).value = c38.value
            
            if 'readColumn39' in globals() and 'writeColumn39' in globals():
                c39 = ws1.cell(row = i, column = readColumn39) 
                ws2.cell(row = i, column = writeColumn39).value = c39.value
            
            if 'readColumn40' in globals() and 'writeColumn40' in globals():
                c40 = ws1.cell(row = i, column = readColumn40) 
                ws2.cell(row = i, column = writeColumn40).value = c40.value
            
            if 'readColumn41' in globals() and 'writeColumn41' in globals():
                c41 = ws1.cell(row = i, column = readColumn41) 
                ws2.cell(row = i, column = writeColumn41).value = c41.value
                
            if 'readColumn42' in globals() and 'writeColumn42' in globals():
                c42 = ws1.cell(row = i, column = readColumn42) 
                ws2.cell(row = i, column = writeColumn42).value = c42.value
            
            if 'readColumn43' in globals() and 'writeColumn343' in globals():
                c43 = ws1.cell(row = i, column = readColumn43) 
                ws2.cell(row = i, column = writeColumn43).value = c43.value
            
            if 'readColumn44' in globals() and 'writeColumn44' in globals():
                c44 = ws1.cell(row = i, column = readColumn44) 
                ws2.cell(row = i, column = writeColumn44).value = c44.value
                
            if 'readColumn45' in globals() and 'writeColumn45' in globals():
                c45 = ws1.cell(row = i, column = readColumn45) 
                ws2.cell(row = i, column = writeColumn32).value = c45.value
                
            
    os.remove(inputXLS)
    wb2.save("blogistica_" + inputXLS)

    
if __name__ == "__main__":
   
    
    fileList = os.listdir("./")
    for fileName in fileList:
        if fileName.endswith(".xls") or fileName.endswith(".XLS"):
           print("Found XLS Files")
           print(fileName)
           p.save_book_as(file_name = fileName, dest_file_name = fileName + ".xlsx")
           os.remove(fileName)
    newXLS = os.listdir("./")
    for xlsxFile in newXLS:
        if xlsxFile.endswith(".xlsx") or xlsxFile.endswith(".XLSX"):
           print("Converted...")
           print(xlsxFile)
           ExcelActivities(xlsxFile)
           
    #WriteExcelOutput()  
    print("Clean Up: ...")
    CleanUp() 